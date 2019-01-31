import gspread
from oauth2client.service_account import ServiceAccountCredentials
import numpy as np
import ast
from tsfel.tsfel.utils.read_json import compute_dictionary
from tsfel.tsfel.utils.eval import compute_complexity
import openpyxl as xl
import numpy as np
import read_json as rj
import ast

def filter_features(dic, filters):
    features_all = list(np.concatenate([list(dic[dk].keys()) for dk in sorted(dic.keys())]))
    list_shown, feat_shown = list(dic.keys()), features_all
    cost_shown = features_all
    if filters['2'] != {}:
        list_hidden = filters['2']['hiddenValues']
        list_shown = [dk for dk in dic.keys() if dk not in list_hidden]
    if filters['1'] != {}:
        feat_hidden = filters['1']['hiddenValues']
        feat_shown = [ff for ff in features_all if ff not in feat_hidden]
    if filters['3'] != {}:
        cost_numbers = filters['3']['hiddenValues']
        cost_hidden = list(np.concatenate([['Constant','Log'] if int(cn) == 1 else ['Squared','Nlog'] if int(cn) == 3 else ['Linear'] if int(cn) == 2 else ['Unknown'] for cn in cost_numbers]))
        cost_shown = []
        for dk in dic.keys():
            cost_shown += [ff for ff in dic[dk].keys() if dic[dk][ff]['Complexity'] not in cost_hidden]
    features_filtered = list(np.concatenate([list(dic[dk].keys()) for dk in sorted(dic.keys()) if dk in list_shown]))
    features_filtered = [ff for ff in features_filtered if ff in feat_shown]
    features_filtered = [cc for cc in features_filtered if cc in cost_shown]

    return features_filtered

def gSheet_interaction(gSheetName):
    FEATURES_JSON = 'tsfel/tsfel/utils/features.json'
    DEFAULT = {'use': 'yes', 'metric': 'euclidean', 'free parameters': '', 'number of features': 1, 'parameters': ''}
    DICTIONARY = compute_dictionary(FEATURES_JSON, DEFAULT)
    scope = ['https://spreadsheets.google.com/feeds',
             'https://www.googleapis.com/auth/drive']
    creds = ServiceAccountCredentials.from_json_keyfile_name('tsfel/tsfel/utils/client_secret.json', scope)
    client = gspread.authorize(creds)
    confManager = client.open(gSheetName)
    sheet = confManager.sheet1
    metadata = confManager.fetch_sheet_metadata()
    list_of_features = sheet.col_values(2)[4:]
    filters = metadata['sheets'][sheet.id]['basicFilter']['criteria']
    list_filt_features = filter_features(DICTIONARY,filters)

    use_or_not = ['TRUE' if lf in list_filt_features else 'FALSE' for lf in list_of_features]

    len_stat = len(DICTIONARY['Statistical'].keys())
    len_temp = len(DICTIONARY['Temporal'].keys())
    len_spec = len(DICTIONARY['Spectral'].keys())

    assert len(list_of_features) <= (len_spec + len_stat + len_temp), \
    "To insert a new feature, please add it to data/features.json with the code in src/utils/features.py"

    # add new feature
    if len(list_of_features) < (len_spec + len_stat + len_temp):
        # new feature was added
        for domain in DICTIONARY.keys():
            for feat in DICTIONARY[domain].keys():
                if feat not in list_of_features:
                    feat_dict = DICTIONARY[domain][feat]
                    param = ''
                    if feat_dict['free parameters']:
                        param = str({"nbins": [10], 'r': [1]})
                    if feat_dict['parameters'] == 'FS':
                        param = str({"fs": 100})
                    curve = feat_dict['Complexity']
                    curves_all = ['Linear','Log','Square','Nlog','Constant']
                    complexity = compute_complexity(feat, domain) if curve not in curves_all else 1 if curve in ['Constant','Log'] else 2 if curve == 'Linear' else 3 
                    new_feat = ['', feat, domain, complexity, param,
                                feat_dict['description']]
                    idx_row = sheet.findall(domain)[-1].row
                    sheet.insert_row(new_feat, idx_row)
        list_of_features = sheet.col_values(2)[4:]
        filters = metadata['sheets'][sheet.id]['basicFilter']['criteria']
        list_filt_features = filter_features(DICTIONARY, filters)
        use_or_not = ['TRUE' if lf in list_filt_features else 'FALSE' for lf in list_of_features]
        
        len_stat = len(DICTIONARY['Statistical'].keys())
        len_temp = len(DICTIONARY['Temporal'].keys())
        len_spec = len(DICTIONARY['Spectral'].keys())
    
    assert 'TRUE' in use_or_not, 'Please select a feature to extract!' + '\n'

    for i in range(len_stat):
        if use_or_not[i] == 'TRUE':
            DICTIONARY['Statistical'][list_of_features[i]]['use'] = 'yes'
            if list_of_features[i] == 'Histogram':
                val = sheet.cell(i + 5, 5).value
                DICTIONARY['Statistical'][list_of_features[i]]['free parameters'] = {
                    'nbins': [ast.literal_eval(val)['nbins']], "r": [ast.literal_eval(val)['r']]}
        else:
            DICTIONARY['Statistical'][list_of_features[i]]['use']='no'



    for i in range(len_temp):
        if use_or_not[i+len_stat] == 'TRUE':
            DICTIONARY['Temporal'][list_of_features[i+len_stat]]['use'] = 'yes'
        else:
            DICTIONARY['Temporal'][list_of_features[i+len_stat]]['use']='no'


    for i in range(len_spec):
        if use_or_not[i+len_stat+len_temp] == 'TRUE':
            val = sheet.cell(i+len_stat+len_temp+5, 5).value
            if not val:
                DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['parameters'] = ''
            else:
                DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['parameters'] = str(ast.literal_eval(val)['fs'])
                DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['use'] = 'yes'
        else:
            DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['use']='no'

    return DICTIONARY

#extract_sheet()

def excel_interaction():
    wb = xl.load_workbook('../../data/Configuration Manager.xlsx')
    sheet = wb['Complete']
    FEATURES_JSON = '../../data/features.json'
    DEFAULT = {'use': 'yes', 'metric': 'euclidean', 'free parameters': '', 'number of features': 1, 'parameters': ''}
    DICTIONARY = rj.compute_dictionary(FEATURES_JSON, DEFAULT)

    list_of_features = [str(i.value) for i in sheet['B']][4:-6]
    use_or_not = [True for i in list_of_features]
    filter = sheet.auto_filter.filterColumn
    if filter:
        init_feat_idx = str(filter[0]).index('filter=[')
        end_feat_idx = str(filter[0])[init_feat_idx:].index(']')
        feat = str(filter[0])[init_feat_idx+7:init_feat_idx+end_feat_idx].split(",")
        _f = []
        for f in feat:
            _f.append(f[3:-1])

        use_or_not = []
        KEYS = ['Statistical', 'Temporal', 'Spectral']
        if _f[0] in DICTIONARY.keys():
            for _key in KEYS:
                for count in DICTIONARY[_key].keys():
                    use_or_not.append(True if str(_key) in _f else False)
        else:
            use_or_not = [True if i in _f else False for i in list_of_features]

    len_stat = len(DICTIONARY['Statistical'].keys())
    len_temp = len(DICTIONARY['Temporal'].keys())
    len_spec = len(DICTIONARY['Spectral'].keys())

    for i in range(len_stat):
        if use_or_not[i]:
            if list_of_features[i] == 'Histogram':
                val = sheet['E' + str(5+i)].value
                DICTIONARY['Statistical'][list_of_features[i]]['free parameters'] = {'nbins': [ast.literal_eval(val)['nbins']], "r": [ast.literal_eval(val)['r']]}
            DICTIONARY['Statistical'][list_of_features[i]]['use'] = 'yes'
            print list_of_features[i]

    for i in range(len_temp):
        if use_or_not[i+len_stat]:
            DICTIONARY['Temporal'][list_of_features[i+len_stat]]['use'] = 'yes'
            print list_of_features[i+len_stat]

    for i in range(len_spec):
        if use_or_not[i+len_stat+len_temp]:
            val = sheet['E' + str(5 + i+len_stat+len_temp)].value
            if not val:
                DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['parameters'] = ''
            else:
                DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['parameters'] = str(ast.literal_eval(val)['fs'])
            DICTIONARY['Spectral'][list_of_features[i+len_stat+len_temp]]['use'] = 'yes'
            print list_of_features[i+len_stat+len_temp]

    return DICTIONARY


def extract_sheet(type_interaction, gSheetName = ""):
    if type_interaction == "excel":
        DICTIONARY = excel_interaction()
    else: # google sheet
        DICTIONARY = gSheet_interaction(gSheetName)

    return DICTIONARY

